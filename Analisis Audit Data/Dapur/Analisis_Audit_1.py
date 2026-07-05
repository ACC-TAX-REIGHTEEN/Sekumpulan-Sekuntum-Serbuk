import datetime
import glob
import os
import re
import configparser
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

print("--> Memulai pemrosesan data audit analytics ...")

config_file = "config.conf"
if not os.path.exists(config_file):
    with open(config_file, "w") as f:
        f.write("[FILTER]\nSHELL\nIRC\nZN\nGT\nFILTER\nJIMCO\nLAIN\nTOP 1\nOLI\n")

config = configparser.ConfigParser(allow_no_value=True)
config.read(config_file)
filter_produk = [str(key).upper() for key in config["FILTER"]]
if "LAIN" not in filter_produk:
    filter_produk.append("LAIN")

ptm_files = glob.glob("PTM*.xlsx")
if not ptm_files:
    print("--> File PTM*.xlsx tidak ditemukan.")
    exit()

ptm_file = ptm_files[0]
xl = pd.ExcelFile(ptm_file)

def sort_key_sheet(sheet_name):
    try:
        return datetime.datetime.strptime(sheet_name, "%m%y")
    except ValueError:
        return datetime.datetime.min

valid_sheets = [s for s in xl.sheet_names if re.match(r"^\d{4}$", s)]
valid_sheets.sort(key=sort_key_sheet)

if not valid_sheets:
    print("--> Tidak ada sheet bulanan (format MMYY) yang valid di file PTM.")
    exit()

id_months = {1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"Mei", 6:"Jun", 
             7:"Jul", 8:"Agu", 9:"Sep", 10:"Okt", 11:"Nov", 12:"Des"}

def clean_key(s):
    if s is None: return ""
    return re.sub(r"[^a-zA-Z0-9]", "", str(s)).lower()

font_title = Font(name="Calibri", size=14, bold=True, color="1F497D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

font_section = Font(name="Calibri", size=11, bold=True, color="1F497D")
fill_section = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
fill_anomaly = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

thin_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
font_bold = Font(name="Calibri", size=11, bold=True)

monthly_macro_data = []
all_detailed_transactions = []

sisa_piutang_col = clean_key("Sisa Piutang")
nilai_faktur_col = clean_key("Nilai Faktur")
umur_japo_col = clean_key("Umur AR base on Tgl Faktur")
nama_penjual_col = clean_key("Nama Penjual")
nama_pelanggan_col = clean_key("Nama Pelanggan")
negara_pelanggan_col = clean_key("Negara Pelanggan")
kontak_pelanggan_col = clean_key("Kontak Pelanggan")
no_faktur_col = clean_key("No. Faktur")

wb_new = openpyxl.Workbook()
wb_new.remove(wb_new.active)

for sheet in valid_sheets:
    dt = datetime.datetime.strptime(sheet, "%m%y")
    col_label = f"{id_months[dt.month]}-{sheet[2:]}"
    
    df_all = pd.read_excel(ptm_file, sheet_name=sheet, header=None)
    header_idx = 0
    for idx, row in df_all.iterrows():
        row_clean = [clean_key(x) for x in row.dropna()]
        if sisa_piutang_col in row_clean or nama_penjual_col in row_clean:
            header_idx = idx
            break

    df = pd.read_excel(ptm_file, sheet_name=sheet, skiprows=header_idx)
    df_raw_for_loop = df.copy()
    df.columns = [clean_key(c) for c in df.columns]

    if not {sisa_piutang_col, umur_japo_col, nama_penjual_col, nama_pelanggan_col}.issubset(df.columns):
        continue

    df["days"] = df[umur_japo_col].astype(str).str.extract(r"(-?\d+)").astype(float).fillna(0)
    
    neg_p = df[negara_pelanggan_col].fillna("").astype(str).str.strip() if negara_pelanggan_col in df.columns else pd.Series([""]*len(df))
    nam_p = df[nama_pelanggan_col].fillna("").astype(str).str.strip()
    df["cust_id"] = np.where(neg_p.str.lower().isin(["", "nan", "none", "nat"]), nam_p, neg_p)

    not_fraud = ~df[nama_penjual_col].astype(str).str.contains("FRAUD", case=False, na=False)
    is_60_364 = df["days"].between(60, 364)
    is_365_up = df["days"] >= 365

    tot_ar = float(pd.to_numeric(df[sisa_piutang_col], errors="coerce").sum())
    ar_60 = float(pd.to_numeric(df[is_60_364 & not_fraud][sisa_piutang_col], errors="coerce").sum())
    ar_365 = float(pd.to_numeric(df[is_365_up & not_fraud][sisa_piutang_col], errors="coerce").sum())
    
    cust_tot = int(df["cust_id"].nunique(dropna=True))
    cust_60 = int(df[is_60_364 & not_fraud]["cust_id"].nunique(dropna=True))
    cust_365 = int(df[is_365_up & not_fraud]["cust_id"].nunique(dropna=True))

    monthly_macro_data.append({
        "Bulan": col_label,
        "Total_AR_Outstanding": tot_ar,
        "AR_60_Hari_Up": ar_60,
        "AR_Bad_Debt": ar_365,
        "Total_Customer": cust_tot,
        "Cust_AR_60_Hari_Up": cust_60,
        "Cust_AR_Bad_Debt": cust_365
    })

    df_raw_for_loop["Bulan_Sistem"] = col_label
    df_raw_for_loop["Cleaned_Sisa_Piutang"] = pd.to_numeric(df[sisa_piutang_col], errors="coerce").fillna(0)
    df_raw_for_loop["Cleaned_Nilai_Faktur"] = pd.to_numeric(df[nilai_faktur_col], errors="coerce").fillna(0) if nilai_faktur_col in df.columns else df_raw_for_loop["Cleaned_Sisa_Piutang"]
    df_raw_for_loop["Cleaned_Days"] = df["days"]
    df_raw_for_loop["Cleaned_Cust_ID"] = df["cust_id"]
    all_detailed_transactions.append(df_raw_for_loop)

df_micro_master = pd.concat(all_detailed_transactions, ignore_index=True)
df_macro_master = pd.DataFrame(monthly_macro_data)

def format_sheet_table(ws, start_row, header_cols, data_matrix, number_formats=None):
    ws.append(header_cols)
    h_row = ws.max_row
    for c_idx in range(1, len(header_cols) + 1):
        cell = ws.cell(row=h_row, column=c_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    for row_idx, row_data in enumerate(data_matrix):
        ws.append(row_data)
        curr_r = ws.max_row
        for c_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_r, column=c_idx)
            cell.border = thin_border
            
            if number_formats and c_idx in number_formats:
                cell.number_format = number_formats[c_idx]
                if "0.0%" in number_formats[c_idx] or "#,##0" in number_formats[c_idx] or "0.00" in number_formats[c_idx]:
                    cell.alignment = align_right
            elif isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = align_right

print("--> Menyusun Sheet 1: Analisis Horisontal & Vertikal...")
ws1 = wb_new.create_sheet(title="1. Horisontal & Vertikal")
ws1.append(["1. LAPORAN AUDIT ANALISIS HORISONTAL & VERTIKAL RASIO PIUTANG"])
ws1.cell(row=1, column=1).font = font_title
ws1.append(["Fokus Audit: Menilai volatilitas perubahan saldo bulanan (Horisontal) dan komposisi risiko portofolio piutang (Vertikal)."])
ws1.append([])

df_macro_calc = df_macro_master.copy()
df_macro_calc["MoM_Total_AR_Chg%"] = df_macro_calc["Total_AR_Outstanding"].pct_change().fillna(0)
df_macro_calc["Rasio_AR_60_Up%"] = (df_macro_calc["AR_60_Hari_Up"] / df_macro_calc["Total_AR_Outstanding"]).fillna(0)
df_macro_calc["Rasio_Bad_Debt%"] = (df_macro_calc["AR_Bad_Debt"] / df_macro_calc["Total_AR_Outstanding"]).fillna(0)

headers_s1 = [
    "Bulan", "Total AR Outstanding", "Perubahan MoM (%)", 
    "AR 60 Hari Up", "Rasio Kontribusi 60+ (%)", 
    "AR Bad Debt (365+)", "Rasio Kontribusi Bad Debt (%)"
]
matrix_s1 = df_macro_calc[[
    "Bulan", "Total_AR_Outstanding", "MoM_Total_AR_Chg%", 
    "AR_60_Hari_Up", "Rasio_AR_60_Up%", "AR_Bad_Debt", "Rasio_Bad_Debt%"
]].values.tolist()

formats_s1 = {
    2: '#,##0', 3: '0.00%;-0.00%;"-"', 
    4: '#,##0', 5: '0.00%;-0.00%;"-"', 
    6: '#,##0', 7: '0.00%;-0.00%;"-"'
}
format_sheet_table(ws1, 4, headers_s1, matrix_s1, formats_s1)

print("--> Menyusun Sheet 3: Deteksi Statistik Z-Score...")
ws2 = wb_new.create_sheet(title="2. Anomali Z-Score Makro")
ws2.append(["2. LAPORAN DETEKSI ANOMALI STATISTIK (METODE Z-SCORE)"])
ws2.cell(row=1, column=1).font = font_title
ws2.append(["Fokus Audit: Menemukan lonjakan saldo bulanan ekstrim yang berada di luar batas fluktuasi normal bisnis secara statistik."])
ws2.append([])

mean_ar = df_macro_master["Total_AR_Outstanding"].mean()
std_ar = df_macro_master["Total_AR_Outstanding"].std()
if std_ar == 0 or pd.isna(std_ar): std_ar = 1

df_macro_calc["Z_Score"] = (df_macro_calc["Total_AR_Outstanding"] - mean_ar) / std_ar
df_macro_calc["Status_Audit"] = np.where(df_macro_calc["Z_Score"].abs() > 1.2, "ANOMALI / REVIU MENDALAM", "WAJAR")

headers_s2 = ["Bulan", "Total AR Outstanding", "Deviasi Nilai dari Rata-Rata (Z-Score)", "Status Kesimpulan Audit"]
matrix_s2 = df_macro_calc[["Bulan", "Total_AR_Outstanding", "Z_Score", "Status_Audit"]].values.tolist()
formats_s2 = {2: '#,##0', 3: '0.00;(-0.00);"-"'}

format_sheet_table(ws2, 4, headers_s2, matrix_s2, formats_s2)

for r_idx in range(5, ws2.max_row + 1):
    status_cell = ws2.cell(row=r_idx, column=4).value
    if status_cell == "ANOMALI / REVIU MENDALAM":
        for c_idx in range(1, 5):
            ws2.cell(row=r_idx, column=c_idx).fill = fill_anomaly
            ws2.cell(row=r_idx, column=c_idx).font = font_bold

print("--> Menyusun Sheet 3: Forensik Digit Angka Hukum Benford...")
ws3 = wb_new.create_sheet(title="3. Forensik Hukum Benford")
ws3.append(["3. ANALISIS FORENSIK DIGIT ANGKA (HUKUM BENFORD)"])
ws3.cell(row=1, column=1).font = font_title
ws3.append(["Fokus Audit: Menguji deviasi frekuensi kemunculan angka pertama faktur. Lonjakan tidak wajar mengindikasikan fraud/faktur fiktif."])
ws3.append([])

def get_first_digit(val):
    val_str = re.sub(r"[^0-9]", "", str(val))
    for char in val_str:
        if char in "123456789":
            return int(char)
    return np.nan

df_micro_master["First_Digit"] = df_micro_master["Cleaned_Nilai_Faktur"].apply(get_first_digit)
df_benford_valid = df_micro_master[df_micro_master["First_Digit"].notna()]
total_benford_count = len(df_benford_valid)

benford_dist = {1: 0.301, 2: 0.176, 3: 0.125, 4: 0.097, 5: 0.079, 6: 0.067, 7: 0.058, 8: 0.051, 9: 0.046}
matrix_s3 = []

if total_benford_count > 0:
    actual_counts = df_benford_valid["First_Digit"].value_counts(normalize=True).to_dict()
    for d in range(1, 10):
        act_p = actual_counts.get(d, 0.0)
        exp_p = benford_dist[d]
        deviasi = act_p - exp_p
        kesimpulan = "REVIU (Deviasi Tinggi)" if abs(deviasi) > 0.05 else "NORMAL"
        matrix_s3.append([f"Digit {d}", act_p, exp_p, deviasi, kesimpulan])
else:
    matrix_s3 = [[f"Digit {d}", 0.0, benford_dist[d], -benford_dist[d], "DATA TIDAK CUKUP"] for d in range(1, 10)]

headers_s3 = ["Komponen Digit Pertama", "Proporsi Riil Faktur (%)", "Ekspektasi Hukum Benford (%)", "Nilai Deviasi Selisih", "Status Evaluasi Data"]
formats_s3 = {2: '0.00%', 3: '0.00%', 4: '0.00%'}

ws3.append(headers_s3)
h_row3 = ws3.max_row
for c_idx in range(1, len(headers_s3) + 1):
    ws3.cell(row=h_row3, column=c_idx).font = font_header
    ws3.cell(row=h_row3, column=c_idx).fill = fill_header
    ws3.cell(row=h_row3, column=c_idx).border = thin_border
    ws3.cell(row=h_row3, column=c_idx).alignment = align_center

for row_data in matrix_s3:
    ws3.append(row_data)
    curr_r = ws3.max_row
    needs_review = row_data[4] == "REVIU (Deviasi Tinggi)"
    for c_idx in range(1, len(row_data) + 1):
        cell = ws3.cell(row=curr_r, column=c_idx)
        cell.border = thin_border
        if c_idx in [2, 3, 4]:
            cell.number_format = '0.00%'
            cell.alignment = align_right
        if needs_review:
            cell.fill = fill_anomaly
            cell.font = font_bold

print("--> Menyusun Sheet 4: Kosentrasi Risiko Piutang Hukum Pareto...")
ws4 = wb_new.create_sheet(title="4. Risiko Pareto Customer")
ws4.append(["4. ANALISIS KONSENTRASI RISIKO PIUTANG (HUKUM PARETO 80/20)"])
ws4.cell(row=1, column=1).font = font_title
ws4.append(["Fokus Audit: Mengidentifikasi segelintir pelanggan inti yang menguasai akumulasi 80% piutang terbesar untuk menilai risiko gagal bayar masal."])
ws4.append([])

latest_month = df_macro_master["Bulan"].iloc[-1]
df_latest_micro = df_micro_master[df_micro_master["Bulan_Sistem"] == latest_month]

df_cust_risk = df_latest_micro.groupby("Cleaned_Cust_ID")["Cleaned_Sisa_Piutang"].sum().reset_index()
df_cust_risk = df_cust_risk.sort_values(by="Cleaned_Cust_ID", ascending=False)
df_cust_risk = df_cust_risk.sort_values(by="Cleaned_Sisa_Piutang", ascending=False)

total_latest_ar = df_cust_risk["Cleaned_Sisa_Piutang"].sum()
if total_latest_ar == 0: total_latest_ar = 1

df_cust_risk["Kontribusi%"] = df_cust_risk["Cleaned_Sisa_Piutang"] / total_latest_ar
df_cust_risk["Kumulatif%"] = df_cust_risk["Kontribusi%"].cumsum()
df_cust_risk["Klasifikasi_Risiko"] = np.where(df_cust_risk["Kumulatif%"] <= 0.80, "TOP 80% CORE RISK (Kritis)", "Regular")

top_25_cust = df_cust_risk.head(25).values.tolist()

headers_s4 = ["ID/Nama Pelanggan", f"Total Saldo Piutang ({latest_month})", "Proporsi Beban Risiko (%)", "Akumulasi Distribusi Kumulatif (%)", "Grup Prioritas Pengawasan"]
formats_s4 = {2: '#,##0', 3: '0.00%', 4: '0.00%'}

ws4.append(headers_s4)
h_row4 = ws4.max_row
for c_idx in range(1, len(headers_s4) + 1):
    ws4.cell(row=h_row4, column=c_idx).font = font_header
    ws4.cell(row=h_row4, column=c_idx).fill = fill_header
    ws4.cell(row=h_row4, column=c_idx).border = thin_border
    ws4.cell(row=h_row4, column=c_idx).alignment = align_center

for row_data in top_25_cust:
    ws4.append(row_data)
    curr_r = ws4.max_row
    is_critical = "Kritis" in row_data[4]
    for c_idx in range(1, len(row_data) + 1):
        cell = ws4.cell(row=curr_r, column=c_idx)
        cell.border = thin_border
        if c_idx == 2:
            cell.number_format = '#,##0'
            cell.alignment = align_right
        elif c_idx in [3, 4]:
            cell.number_format = '0.00%'
            cell.alignment = align_right
        if is_critical:
            cell.fill = fill_anomaly
            cell.font = font_bold

print("--> Menyusun Sheet 5: Uji Duplikasi Nomor Faktur...")
ws5 = wb_new.create_sheet(title="5. Uji Duplikasi Faktur")
ws5.append(["5. LAPORAN AUDIT DETEKSI DUPLIKASI NOMOR FAKTUR (DUPLICATE INVOICE TEST)"])
ws5.cell(row=1, column=1).font = font_title
ws5.append(["Fokus Audit: Mengidentifikasi nomor faktur ganda yang muncul dalam satu bulan transaksi yang sama (Mendeteksi Double Input / Double Journal)."])
ws5.append([])

df_micro_clean = df_micro_master.copy()
df_micro_clean.columns = [clean_key(c) for c in df_micro_clean.columns]

if no_faktur_col in df_micro_clean.columns:
    df_faktur_valid = df_micro_clean[df_micro_clean[no_faktur_col].notna() & (df_micro_clean[no_faktur_col].astype(str).str.strip() != "")]
    dup_counts = df_faktur_valid.groupby(['bulansistem', no_faktur_col]).size().reset_index(name='Count')
    dup_recap = dup_counts[dup_counts['Count'] > 1]
    dup_recap = dup_recap.sort_values(by='Count', ascending=False)
    df_dup_details = pd.merge(df_faktur_valid, dup_recap, on=['bulansistem', no_faktur_col], how='inner')
    
    headers_s5 = ["Nomor Faktur", "Jumlah Kemunculan Di Bulan Ini", "ID/Nama Pelanggan", "Bulan Transaksi", "Nilai Faktur", "Sisa Piutang", "Rekomendasi Tindakan"]
    
    matrix_s5 = []
    for _, row in df_dup_details.iterrows():
        matrix_s5.append([
            str(row[no_faktur_col]),
            int(row['Count']),
            str(row[nama_pelanggan_col] if nama_pelanggan_col in row else row['cleanedcustid']),
            str(row['bulansistem']),
            float(row['cleanednilaifaktur']),
            float(row['cleanedsisapiutang']),
            "REVIU JURNAL (Duplikasi Intra-Bulan)"
        ])
else:
    headers_s5 = ["Status Pengujian"]
    matrix_s5 = [["Kolom Nomor Faktur tidak ditemukan di dalam berkas data sumber."]]

formats_s5 = {2: '#,##0', 5: '#,##0', 6: '#,##0'}
format_sheet_table(ws5, 4, headers_s5, matrix_s5, formats_s5)

print("--> Menyelaraskan ukuran lebar kolom seluruh berkas audit otomatis...")
for sheet_name in wb_new.sheetnames:
    ws_to_fit = wb_new[sheet_name]
    for col in ws_to_fit.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws_to_fit.column_dimensions[col_letter].width = max(max_len + 3, 12)

final_output = "Laporan_Analisis_Prosedur_Audit.xlsx"
wb_new.save(final_output)
print(f"--> SUKSES! Berkas laporan analitis audit '{final_output}' telah diterbitkan.")
