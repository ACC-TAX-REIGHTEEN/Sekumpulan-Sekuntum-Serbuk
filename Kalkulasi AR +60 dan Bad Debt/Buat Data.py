import datetime
import glob
import os
import re
import configparser
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

print("--> Memulai pemrosesan data ...")

config_file = "config.conf"
if not os.path.exists(config_file):
    with open(config_file, "w") as f:
        f.write("[FILTER]\nSHELL\nIRC\nZN\nGT\nFILTER\nJIMCO\nLAIN\nTOP 1\nOLI\n")

config = configparser.ConfigParser(allow_no_value=True)
config.read(config_file)
filter_produk = [str(key).upper() for key in config["FILTER"]]

if "LAIN" not in filter_produk:
    filter_produk.append("LAIN")

ptm_files = glob.glob("PTM*.xlsx")
if not ptm_files:
    print("--> File PTM*.xlsx tidak ditemukan.")
    exit()

ptm_file = ptm_files[0]
xl = pd.ExcelFile(ptm_file)

def sort_key_sheet(sheet_name):
    try:
        return datetime.datetime.strptime(sheet_name, "%m%y")
    except ValueError:
        return datetime.datetime.min

valid_sheets = [s for s in xl.sheet_names if re.match(r"^\d{4}$", s)]
valid_sheets.sort(key=sort_key_sheet)

if not valid_sheets:
    print("--> Tidak ada sheet bulanan (format MMYY) yang valid di file PTM.")
    exit()

id_months = {1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"Mei", 6:"Jun", 
             7:"Jul", 8:"Agu", 9:"Sep", 10:"Okt", 11:"Nov", 12:"Des"}

def clean_key(s):
    if s is None: return ""
    return re.sub(r"[^a-zA-Z0-9]", "", str(s)).lower()

font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

font_section = Font(name="Calibri", size=11, bold=True, color="1F497D")
fill_section = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

thin_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

font_bold = Font(name="Calibri", size=11, bold=True)

def append_formatted_row(ws, row_data, row_type, len_m, len_q):
    ws.append(row_data)
    r = ws.max_row
    
    m_start = 3
    m_end = 2 + len_m
    q_start = m_end + 2
    q_end = q_start + len_q - 1
    d_start = q_end + 2
    d_end = d_start + len_q - 1
    
    is_total_or_selisih = str(row_data[0]).strip() in ["Total", "Selisih"]
    
    for c in range(1, len(row_data) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        
        if is_total_or_selisih:
            cell.font = font_bold
            
        if cell.value is None or cell.value == "":
            continue
        if isinstance(cell.value, str):
            continue
            
        if m_start <= c <= m_end or q_start <= c <= q_end:
            if row_type == "amount":
                cell.number_format = '#,##0;(#,##0);"-"'
            elif row_type == "percent":
                cell.number_format = '0.00%;-0.00%;"-"'
            elif row_type == "count":
                cell.number_format = '#,##0;(#,##0);"-"'
        elif d_start <= c <= d_end:
            cell.number_format = '0.00%;-0.00%;"-"'

monthly_results = {}
quarter_groups = {} 

sisa_piutang_col = clean_key("Sisa Piutang")
umur_japo_col = clean_key("Umur AR base on Tgl Faktur")
nama_penjual_col = clean_key("Nama Penjual")
nama_pelanggan_col = clean_key("Nama Pelanggan")
negara_pelanggan_col = clean_key("Negara Pelanggan")
kontak_pelanggan_col = clean_key("Kontak Pelanggan")

wb_new = openpyxl.Workbook()
ws_rekap = wb_new.active
ws_rekap.title = "Rekap Monitoring"

for sheet in valid_sheets:
    dt = datetime.datetime.strptime(sheet, "%m%y")
    col_label = f"{id_months[dt.month]}-{sheet[2:]}"
    
    q_num = (dt.month - 1) // 3 + 1
    q_label = f"Q{q_num} {dt.year}"
    if q_label not in quarter_groups:
        quarter_groups[q_label] = []
    quarter_groups[q_label].append(col_label)

    df_all = pd.read_excel(ptm_file, sheet_name=sheet, header=None)
    header_idx = 0
    for idx, row in df_all.iterrows():
        row_clean = [clean_key(x) for x in row.dropna()]
        if sisa_piutang_col in row_clean or nama_penjual_col in row_clean:
            header_idx = idx
            break

    df = pd.read_excel(ptm_file, sheet_name=sheet, skiprows=header_idx)
    df_out = df.copy()
    df.columns = [clean_key(c) for c in df.columns]

    if not {sisa_piutang_col, umur_japo_col, nama_penjual_col, nama_pelanggan_col, negara_pelanggan_col}.issubset(df.columns):
        continue

    df["days"] = df[umur_japo_col].astype(str).str.extract(r"(-?\d+)").astype(float)
    
    neg_pel_series = df[negara_pelanggan_col].fillna("").astype(str).str.strip()
    nam_pel_series = df[nama_pelanggan_col].fillna("").astype(str).str.strip()
    is_empty_neg = neg_pel_series.str.lower().isin(["", "nan", "none", "nat"])
    df["cust_id"] = neg_pel_series
    df.loc[is_empty_neg, "cust_id"] = nam_pel_series.loc[is_empty_neg]
    df.loc[df["cust_id"].astype(str).str.strip().str.lower().isin(["", "nan", "none", "nat"]), "cust_id"] = np.nan

    not_fraud = ~df[nama_penjual_col].astype(str).str.contains("FRAUD", case=False, na=False)
    is_60_364 = df["days"].between(60, 364)
    is_365_up = df["days"] >= 365

    tot_ar = float(pd.to_numeric(df[sisa_piutang_col], errors="coerce").sum())
    ar_60 = float(pd.to_numeric(df[is_60_364 & not_fraud][sisa_piutang_col], errors="coerce").sum())
    ar_365 = float(pd.to_numeric(df[is_365_up & not_fraud][sisa_piutang_col], errors="coerce").sum())

    pct_60_amt = float(ar_60 / tot_ar if tot_ar else 0.0)
    pct_365_amt = float(ar_365 / tot_ar if tot_ar else 0.0)

    cust_tot = int(df["cust_id"].nunique(dropna=True))
    cust_60 = int(df[is_60_364 & not_fraud]["cust_id"].nunique(dropna=True))
    cust_365 = int(df[is_365_up & not_fraud]["cust_id"].nunique(dropna=True))

    pct_60_cust = float(cust_60 / cust_tot if cust_tot else 0.0)
    pct_365_cust = float(cust_365 / cust_tot if cust_tot else 0.0)

    prod_60_dict = {p: 0.0 for p in filter_produk}
    prod_365_dict = {p: 0.0 for p in filter_produk}

    if kontak_pelanggan_col in df.columns:
        df["kontak_clean"] = df[kontak_pelanggan_col].fillna("").astype(str).str.upper()
        for _, row_data in df.iterrows():
            piutang_val = pd.to_numeric(row_data[sisa_piutang_col], errors="coerce")
            if pd.isna(piutang_val): continue
            
            matched_prod = None
            for p in filter_produk:
                if p != "LAIN" and p.replace(" ", "") in row_data["kontak_clean"].replace(" ", ""):
                    matched_prod = p
                    break
            if not matched_prod:
                matched_prod = "LAIN"
            
            if row_data["days"] >= 60 and row_data["days"] <= 364 and not "FRAUD" in str(row_data[nama_penjual_col]).upper():
                prod_60_dict[matched_prod] += float(piutang_val)
            elif row_data["days"] >= 365 and not "FRAUD" in str(row_data[nama_penjual_col]).upper():
                prod_365_dict[matched_prod] += float(piutang_val)

    monthly_results[col_label] = {
        "metrics": {
            "TOTAL AR OUTSTANDING (AMOUNT)": tot_ar,
            "AR 60 HARI UP (AMOUNT)": ar_60,
            "AR BAD DEBT (365 UP ) AMOUNT": ar_365,
            "AR 60 HARI UP IN AMOUNT (%)": pct_60_amt,
            "AR BAD DEBT IN AMOUNT (365 UP ) %": pct_365_amt,
            "JUMLAH CUSTOMER (DISTINCT COUNT)": cust_tot,
            "JUMLAH CUST AR 60 HARI UP (DISTINCT COUNT)": cust_60,
            "JUMLAH CUST AR BAD DEBT (365 UP) DISTINCT COUNT": cust_365,
            "AR 60 HARI UP (%)": pct_60_cust,
            "AR BAD DEBT (365 UP ) %": pct_365_cust,
        },
        "prod_60": prod_60_dict,
        "prod_365": prod_365_dict
    }

    df_out["Calculated_Days"] = df["days"]
    df_out["Calculated_Cust_ID"] = df["cust_id"]
    df_out["Is_Fraud"] = ~not_fraud

    df_60 = df_out[is_60_364 & not_fraud]
    df_365 = df_out[is_365_up & not_fraud]
    df_other = df_out[~((is_60_364 & not_fraud) | (is_365_up & not_fraud))]

    new_ws = wb_new.create_sheet(title=f"Detail_{sheet}")
    def write_table(title, sub_df):
        new_ws.append([title])
        new_ws.cell(row=new_ws.max_row, column=1).font = Font(name="Calibri", size=12, bold=True, color="1F497D")
        
        sub_df = sub_df.astype(object).where(pd.notnull(sub_df), None)
        rows_list = list(dataframe_to_rows(sub_df, index=False, header=True))
        
        if rows_list:
            new_ws.append(rows_list[0])
            header_r = new_ws.max_row
            for c in range(1, len(rows_list[0]) + 1):
                cell = new_ws.cell(row=header_r, column=c)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            for r_data in rows_list[1:]:
                new_ws.append(r_data)
                curr_r = new_ws.max_row
                for c in range(1, len(r_data) + 1):
                    new_ws.cell(row=curr_r, column=c).border = thin_border
                    
        new_ws.append([])
        new_ws.append([])

    write_table("BARIS DATA FILTER: AR 60 HARI UP (AMOUNT) & JUMLAH CUST AR 60 HARI UP", df_60)
    write_table("BARIS DATA FILTER: AR BAD DEBT (365 UP) AMOUNT & JUMLAH CUST AR BAD DEBT", df_365)
    write_table("BARIS DATA LAINNYA (TIDAK MASUK ANALISIS AR 60+ ATAU BAD DEBT)", df_other)

months_list = list(monthly_results.keys())
quarters_list = list(quarter_groups.keys())

header_row = ["KETERANGAN", "Formula"] + months_list + [""] + quarters_list + [""]
target_metrics_decline = [
    "AR 60 HARI UP (AMOUNT)",
    "AR BAD DEBT (365 UP ) AMOUNT",
    "JUMLAH CUST AR 60 HARI UP (DISTINCT COUNT)",
    "JUMLAH CUST AR BAD DEBT (365 UP) DISTINCT COUNT"
]

oldest_q = quarters_list[0] if quarters_list else ""
for q in quarters_list:
    header_row.append(f"Penurunan vs {oldest_q} ({q})")

ws_rekap.append([])
ws_rekap.append([])
ws_rekap.append(header_row)
header_row_idx = ws_rekap.max_row

for c_idx in range(1, len(header_row) + 1):
    cell = ws_rekap.cell(row=header_row_idx, column=c_idx)
    if cell.value != "":
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

main_structure = [
    {"label": "TOTAL AR OUTSTANDING (AMOUNT)", "formula": "A"},
    {"label": "AR 60 HARI UP (AMOUNT)", "formula": "B"},
    {"label": "AR BAD DEBT (365 UP ) AMOUNT", "formula": "C"},
    {"label": "SPACER"},
    {"label": "AR 60 HARI UP IN AMOUNT (%)", "formula": "B : A"},
    {"label": "AR BAD DEBT IN AMOUNT (365 UP ) %", "formula": "C : A"},
    {"label": "SPACER"},
    {"label": "JUMLAH CUSTOMER (DISTINCT COUNT)", "formula": "D"},
    {"label": "JUMLAH CUST AR 60 HARI UP (DISTINCT COUNT)", "formula": "E"},
    {"label": "JUMLAH CUST AR BAD DEBT (365 UP) DISTINCT COUNT", "formula": "F"},
    {"label": "SPACER"},
    {"label": "AR 60 HARI UP (%)", "formula": "E : D"},
    {"label": "AR BAD DEBT (365 UP ) %", "formula": "F : D"},
    {"label": "SPACER"}
]

cached_q_averages = {}

for item in main_structure:
    if item["label"] == "SPACER":
        ws_rekap.append([])
        continue
        
    label = item["label"]
    formula = item["formula"]
    row_data = [label, formula]
    
    for m in months_list:
        row_data.append(monthly_results[m]["metrics"][label])
        
    row_data.append("")
    
    cached_q_averages[label] = {}
    for q in quarters_list:
        count_months = len(quarter_groups[q])
        q_sum = sum(monthly_results[m]["metrics"][label] for m in quarter_groups[q])
        q_avg = q_sum / count_months if count_months > 0 else 0.0
        cached_q_averages[label][q] = q_avg
        row_data.append(q_avg)
            
    row_data.append("")
    
    for q in quarters_list:
        if label in target_metrics_decline and oldest_q in cached_q_averages[label]:
            val_oldest = cached_q_averages[label][oldest_q]
            val_current = cached_q_averages[label][q]
            if val_oldest > 0:
                row_data.append((val_current - val_oldest) / val_oldest)
            else:
                row_data.append(0.0)
        else:
            row_data.append("-")
            
    r_type = "amount"
    if "%" in label or "%)" in label:
        r_type = "percent"
    elif "customer" in label.lower() or "count" in label.lower():
        r_type = "count"
        
    append_formatted_row(ws_rekap, row_data, r_type, len(months_list), len(quarters_list))

def write_horizontal_contrib_table(title, data_key, main_comparison_key):
    ws_rekap.append([title])
    sec_r = ws_rekap.max_row
    ws_rekap.cell(row=sec_r, column=1).font = font_section
    ws_rekap.cell(row=sec_r, column=1).fill = fill_section
    
    for p in filter_produk:
        prod_row = [p, ""]
        for m in months_list:
            prod_row.append(float(monthly_results[m][data_key][p]))
        
        prod_row.append("")
        for q in quarters_list:
            count_months = len(quarter_groups[q])
            q_avg_prod = sum(monthly_results[m][data_key][p] for m in quarter_groups[q]) / count_months if count_months > 0 else 0.0
            prod_row.append(float(q_avg_prod))
        prod_row.append("")
        for q in quarters_list: prod_row.append("-")
        
        append_formatted_row(ws_rekap, prod_row, "amount", len(months_list), len(quarters_list))
        
    total_row = ["Total", ""]
    for m in months_list:
        m_total = sum(monthly_results[m][data_key][p] for p in filter_produk)
        total_row.append(float(m_total))
    
    total_row.append("")
    for q in quarters_list:
        count_months = len(quarter_groups[q])
        q_total_avg = sum(sum(monthly_results[m][data_key][p] for p in filter_produk) for m in quarter_groups[q]) / count_months if count_months > 0 else 0.0
        total_row.append(float(q_total_avg))
    total_row.append("")
    for q in quarters_list: total_row.append("-")
    append_formatted_row(ws_rekap, total_row, "amount", len(months_list), len(quarters_list))
    
    selisih_row = ["Selisih", ""]
    for m in months_list:
        m_total = sum(monthly_results[m][data_key][p] for p in filter_produk)
        main_val = monthly_results[m]["metrics"][main_comparison_key]
        selisih_row.append(float(m_total - main_val))
        
    selisih_row.append("")
    for q in quarters_list:
        count_months = len(quarter_groups[q])
        q_total_avg = sum(sum(monthly_results[m][data_key][p] for p in filter_produk) for m in quarter_groups[q]) / count_months if count_months > 0 else 0.0
        main_q_avg = cached_q_averages[main_comparison_key][q]
        selisih_row.append(float(q_total_avg - main_q_avg))
    selisih_row.append("")
    for q in quarters_list: selisih_row.append("-")
    append_formatted_row(ws_rekap, selisih_row, "amount", len(months_list), len(quarters_list))
    
    ws_rekap.append([])

write_horizontal_contrib_table("Daftar Kontribusi Barang AR-60 HARI", "prod_60", "AR 60 HARI UP (AMOUNT)")
write_horizontal_contrib_table("Daftar Kontribusi Barang BADDEBT", "prod_365", "AR BAD DEBT (365 UP ) AMOUNT")

print("--> Menerapkan autofit lebar kolom untuk seluruh sheet...")
for sheet_name in wb_new.sheetnames:
    ws_to_fit = wb_new[sheet_name]
    for col in ws_to_fit.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws_to_fit.column_dimensions[col_letter].width = max(max_len + 3, 11)

output_file = "Monitoring.xlsx"
wb_new.save(output_file)
print(f"--> Selesai sempurna! File '{output_file}' siap digunakan.")
